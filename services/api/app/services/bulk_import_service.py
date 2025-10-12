from __future__ import annotations

import posixpath
import re
from dataclasses import dataclass, field
from datetime import datetime
from io import BytesIO
from typing import Any, Dict, Iterable, List, Sequence, Tuple
from xml.etree import ElementTree as ET
from zipfile import BadZipFile, ZipFile


class BulkImportFormatError(ValueError):
    """Raised when the uploaded workbook cannot be parsed."""


@dataclass
class ParsedSubject:
    source_row: int
    name: str
    description: str | None
    icon: str | None
    errors: List[str] = field(default_factory=list)


@dataclass
class ParsedQuiz:
    source_row: int
    title: str
    description: str | None
    is_active: bool
    question_prompts: List[str]
    errors: List[str] = field(default_factory=list)


@dataclass
class ParsedQuestionOption:
    text: str
    is_correct: bool = False


@dataclass
class ParsedQuestion:
    source_row: int
    prompt: str
    explanation: str | None
    subject_label: str | None
    difficulty: str | None
    is_active: bool
    subject_name: str
    quiz_titles: List[str]
    options: List[ParsedQuestionOption]
    errors: List[str] = field(default_factory=list)


@dataclass
class ParsedWorkbook:
    subjects: List[ParsedSubject]
    quizzes: List[ParsedQuiz]
    questions: List[ParsedQuestion]
    warnings: List[str] = field(default_factory=list)


@dataclass
class ExportSubject:
    name: str
    description: str | None
    icon: str | None


@dataclass
class ExportQuiz:
    title: str
    description: str | None
    is_active: bool
    question_prompts: List[str]


@dataclass
class ExportQuestionOption:
    text: str
    is_correct: bool


@dataclass
class ExportQuestion:
    prompt: str
    explanation: str | None
    subject_label: str | None
    difficulty: str | None
    is_active: bool
    subject_name: str
    quiz_titles: List[str]
    options: List[ExportQuestionOption]


_SUBJECT_SHEET_NAMES = {"subjects", "subject", "subject setup"}
_QUIZ_SHEET_NAMES = {"quizzes", "quiz", "quiz setup"}
_QUESTION_SHEET_NAMES = {"questions", "question bank", "items"}

_NS = {
    "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "rel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
}


def parse_workbook(file_bytes: bytes) -> ParsedWorkbook:
    try:
        sheet_map = _load_sheet_map(file_bytes)
    except (BadZipFile, KeyError, ET.ParseError) as exc:  # noqa: BLE001
        raise BulkImportFormatError("Unable to read the Excel workbook. Upload a valid .xlsx file.") from exc

    subjects_sheet = _locate_sheet(sheet_map.keys(), _SUBJECT_SHEET_NAMES)
    quizzes_sheet = _locate_sheet(sheet_map.keys(), _QUIZ_SHEET_NAMES)
    questions_sheet = _locate_sheet(sheet_map.keys(), _QUESTION_SHEET_NAMES)

    warnings: List[str] = []
    if subjects_sheet is None:
        warnings.append("Subjects sheet not found. Expected a sheet named 'Subjects'.")
    if quizzes_sheet is None:
        warnings.append("Quizzes sheet not found. Expected a sheet named 'Quizzes'.")
    if questions_sheet is None:
        warnings.append("Questions sheet not found. Expected a sheet named 'Questions'.")

    subjects = _parse_subjects(sheet_map[subjects_sheet]) if subjects_sheet else []
    quizzes = _parse_quizzes(sheet_map[quizzes_sheet]) if quizzes_sheet else []
    questions = _parse_questions(sheet_map[questions_sheet]) if questions_sheet else []

    return ParsedWorkbook(
        subjects=subjects,
        quizzes=quizzes,
        questions=questions,
        warnings=warnings,
    )


def _locate_sheet(sheet_names: Iterable[str], expected: set[str]) -> str | None:
    normalized = {_normalize_sheet_name(name): name for name in sheet_names}
    for candidate in expected:
        candidate_key = _normalize_sheet_name(candidate)
        if candidate_key in normalized:
            return normalized[candidate_key]
    return None


def _normalize_sheet_name(name: str) -> str:
    normalized = name.strip().lower()
    return re.sub(r"[^a-z0-9]+", "", normalized)


def _parse_subjects(rows: List[List[Any]]) -> List[ParsedSubject]:
    headers = _extract_headers(rows)
    subjects: List[ParsedSubject] = []
    for row_idx, values in _iter_rows(rows):
        row_map = _row_to_map(headers, values)
        name = _normalize_str(_pick(row_map, ["name", "subject", "subject name"]))
        description = _normalize_str(_pick(row_map, ["description", "details", "summary"]))
        icon = _normalize_str(_pick(row_map, ["icon", "emoji"]))
        if not name:
            if not _is_empty_row(values):
                subjects.append(
                    ParsedSubject(
                        source_row=row_idx,
                        name="",
                        description=None,
                        icon=None,
                        errors=["Subject name is required."],
                    )
                )
            continue
        subjects.append(
            ParsedSubject(
                source_row=row_idx,
                name=name,
                description=description,
                icon=icon,
            )
        )
    return subjects


def _parse_quizzes(rows: List[List[Any]]) -> List[ParsedQuiz]:
    headers = _extract_headers(rows)
    quizzes: List[ParsedQuiz] = []
    for row_idx, values in _iter_rows(rows):
        row_map = _row_to_map(headers, values)
        title = _normalize_str(_pick(row_map, ["title", "quiz", "name"]))
        description = _normalize_str(_pick(row_map, ["description", "details"]))
        is_active = _parse_bool(_pick(row_map, ["is active", "active", "status"]), default=True)
        question_raw = _pick(row_map, ["questions", "question prompts", "prompt list"])
        question_prompts = _split_list(question_raw)

        if not title:
            if not _is_empty_row(values):
                quizzes.append(
                    ParsedQuiz(
                        source_row=row_idx,
                        title="",
                        description=None,
                        is_active=is_active,
                        question_prompts=[],
                        errors=["Quiz title is required."],
                    )
                )
            continue

        quizzes.append(
            ParsedQuiz(
                source_row=row_idx,
                title=title,
                description=description,
                is_active=is_active,
                question_prompts=question_prompts,
            )
        )
    return quizzes


def _parse_questions(rows: List[List[Any]]) -> List[ParsedQuestion]:
    headers = _extract_headers(rows)
    questions: List[ParsedQuestion] = []
    for row_idx, values in _iter_rows(rows):
        row_map = _row_to_map(headers, values)
        prompt = _normalize_str(_pick(row_map, ["prompt", "question", "text"]))
        explanation = _normalize_str(_pick(row_map, ["explanation", "rationale", "notes"]))
        subject = _normalize_str(_pick(row_map, ["subject", "topic"]))
        difficulty = _normalize_str(_pick(row_map, ["difficulty", "level"]))
        is_active = _parse_bool(_pick(row_map, ["is active", "active", "status"]), default=True)
        subject_name = _normalize_str(_pick(row_map, ["subject", "subject name"]))
        quiz_titles = _split_list(_pick(row_map, ["quizzes", "quiz titles", "assign to quizzes"]))

        option_pairs = _extract_options(headers, values)
        correct_value = _normalize_str(_pick(row_map, ["correct option", "answer", "correct"]))
        options = _resolve_options(option_pairs, correct_value)

        errors: List[str] = []
        if not prompt:
            if _is_empty_row(values):
                continue
            errors.append("Question prompt is required.")
        if not subject_name:
            errors.append("Subject name is required for each question.")
        if len(options) < 2:
            errors.append("Provide at least two options.")
        elif not any(option.is_correct for option in options):
            errors.append("Select a correct option.")

        if not prompt and _is_empty_row(values):
            continue

        questions.append(
            ParsedQuestion(
                source_row=row_idx,
                prompt=prompt or "",
                explanation=explanation,
                subject=subject,
                difficulty=difficulty,
                is_active=is_active,
                subject_name=subject_name or "",
                quiz_titles=quiz_titles,
                options=options,
                errors=errors,
            )
        )
    return questions


def _extract_headers(rows: List[List[Any]]) -> List[str]:
    if not rows:
        return []
    return [_normalize_header(value) for value in rows[0]]


def _iter_rows(rows: List[List[Any]]) -> Iterable[Tuple[int, Tuple[Any, ...]]]:
    for index, row in enumerate(rows[1:], start=2):
        yield index, tuple(row)


def _row_to_map(headers: List[str], values: Tuple[Any, ...]) -> Dict[str, Any]:
    row: Dict[str, Any] = {}
    for index, header in enumerate(headers):
        if not header:
            continue
        value = values[index] if index < len(values) else None
        row[header] = value
    return row


def _extract_options(headers: List[str], values: Tuple[Any, ...]) -> List[Tuple[str, str]]:
    options: List[Tuple[str, str]] = []
    for index, header in enumerate(headers):
        if not header or not header.startswith("option"):
            continue
        value = values[index] if index < len(values) else None
        normalized_value = _normalize_str(value)
        if normalized_value:
            options.append((header, normalized_value))
    return options


def _resolve_options(option_pairs: List[Tuple[str, str]], correct_value: str | None) -> List[ParsedQuestionOption]:
    options: List[ParsedQuestionOption] = []
    correct_index = _resolve_correct_index(option_pairs, correct_value)
    for idx, (_, text) in enumerate(option_pairs):
        options.append(ParsedQuestionOption(text=text, is_correct=(idx == correct_index)))
    return options


def _resolve_correct_index(option_pairs: List[Tuple[str, str]], correct_value: str | None) -> int | None:
    if correct_value is None:
        return None
    normalized = correct_value.lower()
    for idx, (header, text) in enumerate(option_pairs):
        header_key = header.replace("option", "").strip()
        candidates = {
            text.lower(),
            header.lower(),
            header_key.lower(),
            str(idx + 1),
        }
        if idx < 26:
            candidates.add(chr(ord("a") + idx))
        if normalized in candidates:
            return idx
    return None


def _pick(row: Dict[str, Any], keys: List[str]) -> Any:
    for key in keys:
        if key in row:
            return row[key]
    return None


def _split_list(value: Any) -> List[str]:
    text = _normalize_str(value)
    if not text:
        return []
    separators = [",", ";", "|"]
    for separator in separators[1:]:
        text = text.replace(separator, separators[0])
    parts = [part.strip() for part in text.split(separators[0])]
    return [part for part in parts if part]


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().lower()


def _normalize_str(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        trimmed = value.strip()
        return trimmed or None
    converted = str(value).strip()
    return converted or None


def _parse_bool(value: Any, *, default: bool = True) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0
    normalized = str(value).strip().lower()
    if normalized in {"true", "yes", "y", "1", "active", "publish"}:
        return True
    if normalized in {"false", "no", "n", "0", "inactive", "draft"}:
        return False
    return default


def _is_empty_row(values: Tuple[Any, ...]) -> bool:
    return all(_normalize_str(value) is None for value in values)


def _load_sheet_map(file_bytes: bytes) -> Dict[str, List[List[Any]]]:
    try:
        from openpyxl import load_workbook
        from openpyxl.utils.exceptions import InvalidFileException
    except ImportError:
        return _load_sheet_map_from_archive(file_bytes)

    try:
        workbook = load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    except (InvalidFileException, BadZipFile, KeyError, ET.ParseError, OSError) as exc:  # pragma: no cover - openpyxl raises InvalidFileException
        raise BulkImportFormatError("Unable to read the Excel workbook. Upload a valid .xlsx file.") from exc

    try:
        sheet_map: Dict[str, List[List[Any]]] = {}
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
            sheet_map[sheet_name] = rows
        return sheet_map
    finally:
        workbook.close()


def _load_sheet_map_from_archive(file_bytes: bytes) -> Dict[str, List[List[Any]]]:
    with ZipFile(BytesIO(file_bytes)) as archive:
        workbook_xml = archive.read("xl/workbook.xml")
        workbook_tree = ET.fromstring(workbook_xml)
        rels_tree = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
        relationships = {
            rel.attrib["Id"]: rel.attrib["Target"]
            for rel in rels_tree.findall("rel:Relationship", {"rel": _NS["rel"]})
        }

        shared_strings = _parse_shared_strings(archive)

        sheets: Dict[str, List[List[Any]]] = {}
        for sheet in workbook_tree.findall("main:sheets/main:sheet", _NS):
            name = sheet.attrib["name"]
            rel_id = sheet.attrib[f"{{{_NS['rel']}}}id"]
            target = relationships.get(rel_id)
            if not target:
                continue
            sheet_path = _resolve_sheet_path(archive, target)
            sheet_rows = _parse_sheet(archive, sheet_path, shared_strings)
            sheets[name] = sheet_rows
        return sheets


def _resolve_sheet_path(archive: ZipFile, target: str) -> str:
    if target.startswith("/"):
        candidate = target.lstrip("/")
        if candidate in archive.namelist():
            return candidate
        return candidate

    candidate = posixpath.normpath(posixpath.join("xl", target))
    if candidate in archive.namelist():
        return candidate

    alt = target.lstrip("./")
    if alt in archive.namelist():
        return alt

    return candidate


def _parse_sheet(archive: ZipFile, sheet_path: str, shared_strings: List[str]) -> List[List[Any]]:
    try:
        xml = archive.read(sheet_path)
    except KeyError as exc:
        raise BulkImportFormatError(f"Worksheet '{sheet_path}' is missing from the workbook.") from exc

    tree = ET.fromstring(xml)
    rows: List[List[Any]] = []
    for row in tree.findall("main:sheetData/main:row", _NS):
        row_map: Dict[int, Any] = {}
        max_index = 0
        for cell in row.findall("main:c", _NS):
            ref = cell.attrib.get("r")
            column_index = _column_index(ref)
            value = _parse_cell(cell, shared_strings)
            row_map[column_index] = value
            if column_index > max_index:
                max_index = column_index
        if max_index == 0:
            rows.append([])
            continue
        row_values = [row_map.get(index) for index in range(1, max_index + 1)]
        rows.append(row_values)
    return rows


def _parse_shared_strings(archive: ZipFile) -> List[str]:
    try:
        xml = archive.read("xl/sharedStrings.xml")
    except KeyError:
        return []
    tree = ET.fromstring(xml)
    values: List[str] = []
    for item in tree.findall("main:si", _NS):
        texts = [node.text or "" for node in item.findall(".//main:t", _NS)]
        values.append("".join(texts))
    return values


def _parse_cell(cell: ET.Element, shared_strings: List[str]) -> Any:
    cell_type = cell.attrib.get("t")
    if cell_type == "s":
        index_text = cell.findtext("main:v", namespaces=_NS)
        try:
            index = int(index_text) if index_text is not None else 0
        except ValueError:
            index = 0
        if 0 <= index < len(shared_strings):
            return shared_strings[index]
        return ""
    if cell_type == "b":
        value = cell.findtext("main:v", namespaces=_NS)
        return value in {"1", "true", "TRUE"}
    if cell_type == "inlineStr":
        texts = [node.text or "" for node in cell.findall("main:is/main:t", _NS)]
        return "".join(texts)
    value = cell.findtext("main:v", namespaces=_NS)
    return value


def _column_index(cell_ref: str | None) -> int:
    if not cell_ref:
        return 1
    letters = "".join(char for char in cell_ref if char.isalpha())
    if not letters:
        return 1
    index = 0
    for char in letters:
        index = index * 26 + (ord(char.upper()) - ord("A") + 1)
    return index


def build_bulk_import_template() -> bytes:
    sample_subjects = [
        ExportSubject(name="General Knowledge", description="Mixed trivia sample", icon="sparkles"),
    ]
    sample_quizzes = [
        ExportQuiz(
            title="General Quiz",
            description="Starter quiz to demonstrate the format",
            is_active=True,
            question_prompts=["What is 2 + 2?"],
        )
    ]
    sample_questions = [
        ExportQuestion(
            prompt="What is 2 + 2?",
            explanation="Basic arithmetic question.",
            subject="Mathematics",
            difficulty="Easy",
            is_active=True,
            subject_name="General Knowledge",
            quiz_titles=["General Quiz"],
            options=[
                ExportQuestionOption(text="4", is_correct=True),
                ExportQuestionOption(text="5", is_correct=False),
                ExportQuestionOption(text="3", is_correct=False),
                ExportQuestionOption(text="22", is_correct=False),
            ],
        )
    ]
    return build_bulk_import_workbook(sample_subjects, sample_quizzes, sample_questions)


def build_bulk_import_workbook(
    subjects: Sequence[ExportSubject],
    quizzes: Sequence[ExportQuiz],
    questions: Sequence[ExportQuestion],
) -> bytes:
    subjects_rows: List[List[Any]] = [["Name", "Description", "Icon"]]
    for subject in subjects:
        subjects_rows.append([
            subject.name,
            subject.description or "",
            subject.icon or "",
        ])

    quizzes_rows: List[List[Any]] = [["Title", "Description", "Is Active", "Questions"]]
    for quiz in quizzes:
        quizzes_rows.append([
            quiz.title,
            quiz.description or "",
            quiz.is_active,
            ", ".join(quiz.question_prompts),
        ])

    option_width = max((len(question.options) for question in questions), default=0)
    option_width = max(option_width, 2)
    option_headers = [f"Option {index}" for index in range(1, option_width + 1)]
    questions_header = [
        "Prompt",
        "Explanation",
        "Subject Label",
        "Difficulty",
        "Is Active",
        "Subject",
        *option_headers,
        "Correct Option",
        "Quizzes",
    ]

    questions_rows: List[List[Any]] = [questions_header]
    for question in questions:
        row: List[Any] = [
            question.prompt,
            question.explanation or "",
            question.subject_label or "",
            question.difficulty or "",
            question.is_active,
            question.subject_name,
        ]
        for index in range(option_width):
            option = question.options[index] if index < len(question.options) else None
            row.append(option.text if option else "")
        correct_index = next((idx for idx, option in enumerate(question.options) if option.is_correct), None)
        row.append(f"Option {correct_index + 1}" if correct_index is not None else "")
        row.append(", ".join(question.quiz_titles))
        questions_rows.append(row)

    sheets = {
        "Subjects": subjects_rows,
        "Quizzes": quizzes_rows,
        "Questions": questions_rows,
    }
    return _write_workbook(sheets)


def _write_workbook(sheets: Dict[str, List[List[Any]]]) -> bytes:
    try:
        from openpyxl import Workbook
    except ImportError as exc:  # pragma: no cover - dependency should be installed in production
        raise RuntimeError(
            "openpyxl is required to export bulk import workbooks. "
            "Install the 'openpyxl' package to enable downloads."
        ) from exc

    workbook = Workbook()
    # Remove the default sheet so the ordering exactly matches the requested sheets.
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    for name, rows in sheets.items():
        worksheet = workbook.create_sheet(title=name)
        for row in rows:
            worksheet.append(row)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
